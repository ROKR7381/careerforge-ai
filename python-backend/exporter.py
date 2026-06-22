import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from schemas import ResumeData

def export_to_excel(resume: ResumeData) -> bytes:
    """
    Generates a beautifully formatted, multi-sheet Excel file representing the resume.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Common styles
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Navy Blue
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1A365D")
    subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="555555")
    regular_font = Font(name="Segoe UI", size=11, color="333333")
    bold_font = Font(name="Segoe UI", size=11, bold=True, color="111111")
    section_title_font = Font(name="Segoe UI", size=13, bold=True, color="2C3E50")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Helper function to style a table header row
    def style_header(ws, col_start, col_end, row=1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[row].height = 28

    # Helper to auto-fit columns
    def auto_fit_columns(ws, padding=3, min_width=12, max_width=50):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # Handle linebreaks for length
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            col_letter = get_column_letter(col[0].column)
            fit_width = max(max_len + padding, min_width)
            ws.column_dimensions[col_letter].width = min(fit_width, max_width)

    # ================= Sheet 1: Personal Info & Summary =================
    ws1 = wb.create_sheet(title="Summary")
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1["A1"] = resume.personal_info.full_name
    ws1["A1"].font = title_font
    
    ws1["A2"] = resume.personal_info.professional_title
    ws1["A2"].font = subtitle_font
    
    # Contact details Table
    ws1["A4"] = "Contact Detail"
    ws1["A4"].font = section_title_font
    ws1["B4"] = "Value"
    ws1["B4"].font = section_title_font
    
    contacts = [
        ("Email", resume.personal_info.email),
        ("Phone", resume.personal_info.phone),
        ("Location", resume.personal_info.location),
        ("LinkedIn", resume.personal_info.linkedin or "-"),
        ("GitHub", resume.personal_info.github or "-"),
        ("Website/Portfolio", resume.personal_info.website or "-"),
        ("Date / Place of Birth", resume.personal_info.birth_info or "-"),
        ("Nationality", resume.personal_info.nationality or "-"),
        ("Hobbies", resume.personal_info.hobbies or "-"),
    ]
    
    current_row = 5
    for label, val in contacts:
        ws1.cell(row=current_row, column=1, value=label).font = bold_font
        ws1.cell(row=current_row, column=1).border = thin_border
        ws1.cell(row=current_row, column=1).alignment = align_left
        
        ws1.cell(row=current_row, column=2, value=val).font = regular_font
        ws1.cell(row=current_row, column=2).border = thin_border
        ws1.cell(row=current_row, column=2).alignment = align_left
        current_row += 1
        
    # Professional Summary
    ws1.cell(row=current_row + 1, column=1, value="Professional Summary").font = section_title_font
    summary_cell = ws1.cell(row=current_row + 2, column=1, value=resume.summary)
    summary_cell.font = regular_font
    summary_cell.alignment = align_top_left
    # Merge summary across A and B columns
    ws1.merge_cells(start_row=current_row + 2, start_column=1, end_row=current_row + 6, end_column=2)
    
    # Apply border to merged summary cells
    for r in range(current_row + 2, current_row + 7):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = thin_border
            
    ws1.row_dimensions[current_row + 2].height = 20
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 50

    # ================= Sheet 2: Work Experience =================
    ws2 = wb.create_sheet(title="Work Experience")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = ["Company", "Position", "Location", "Start Date", "End Date", "Achievements & Responsibilities"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(headers))
    
    row_idx = 2
    for exp in resume.experience:
        bullets = "\n".join([f"• {b}" for b in exp.description])
        
        c1 = ws2.cell(row=row_idx, column=1, value=exp.company)
        c2 = ws2.cell(row=row_idx, column=2, value=exp.position)
        c3 = ws2.cell(row=row_idx, column=3, value=exp.location)
        c4 = ws2.cell(row=row_idx, column=4, value=exp.start_date)
        c5 = ws2.cell(row=row_idx, column=5, value=exp.end_date)
        c6 = ws2.cell(row=row_idx, column=6, value=bullets)
        
        # Formatting
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_left
        
        c1.font = bold_font # bold company name
        c6.font = regular_font
        c6.border = thin_border
        c6.alignment = align_top_left
        
        # Approximate height based on bullet count
        ws2.row_dimensions[row_idx].height = max(24, len(exp.description) * 18)
        row_idx += 1
        
    auto_fit_columns(ws2, min_width=15, max_width=65)

    # ================= Sheet 3: Education =================
    ws3 = wb.create_sheet(title="Education")
    ws3.views.sheetView[0].showGridLines = True
    
    headers = ["Institution", "Degree", "Location", "Start Date", "End Date", "Details"]
    for i, h in enumerate(headers, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, 1, len(headers))
    
    row_idx = 2
    for edu in resume.education:
        c1 = ws3.cell(row=row_idx, column=1, value=edu.institution)
        c2 = ws3.cell(row=row_idx, column=2, value=edu.degree)
        c3 = ws3.cell(row=row_idx, column=3, value=edu.location)
        c4 = ws3.cell(row=row_idx, column=4, value=edu.start_date)
        c5 = ws3.cell(row=row_idx, column=5, value=edu.end_date)
        c6 = ws3.cell(row=row_idx, column=6, value=edu.description or "")
        
        for cell in [c1, c2, c3, c4, c5, c6]:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_left
            
        c1.font = bold_font
        ws3.row_dimensions[row_idx].height = 24
        row_idx += 1
        
    auto_fit_columns(ws3, min_width=15, max_width=50)

    # ================= Sheet 4: Skills =================
    ws4 = wb.create_sheet(title="Skills")
    ws4.views.sheetView[0].showGridLines = True
    
    headers = ["Skill Category", "Skills List"]
    for i, h in enumerate(headers, 1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, 1, len(headers))
    
    row_idx = 2
    for sk in resume.skills:
        skills_str = ", ".join(sk.skills)
        c1 = ws4.cell(row=row_idx, column=1, value=sk.category)
        c2 = ws4.cell(row=row_idx, column=2, value=skills_str)
        
        c1.font = bold_font
        c1.border = thin_border
        c1.alignment = align_left
        
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = align_top_left
        
        ws4.row_dimensions[row_idx].height = 24
        row_idx += 1
        
    auto_fit_columns(ws4, min_width=20, max_width=60)

    # ================= Sheet 5: Projects & Certs =================
    ws5 = wb.create_sheet(title="Projects & Certs")
    ws5.views.sheetView[0].showGridLines = True
    
    # Projects Section
    ws5["A1"] = "Projects"
    ws5["A1"].font = section_title_font
    
    headers_proj = ["Project Name", "Role", "Link", "Description"]
    for i, h in enumerate(headers_proj, 1):
        ws5.cell(row=2, column=i, value=h)
    style_header(ws5, 1, len(headers_proj), row=2)
    
    row_idx = 3
    for proj in resume.projects:
        c1 = ws5.cell(row=row_idx, column=1, value=proj.name)
        c2 = ws5.cell(row=row_idx, column=2, value=proj.role or "")
        c3 = ws5.cell(row=row_idx, column=3, value=proj.link or "")
        c4 = ws5.cell(row=row_idx, column=4, value=proj.description)
        
        for cell in [c1, c2, c3, c4]:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_left
            
        c1.font = bold_font
        ws5.row_dimensions[row_idx].height = 35
        row_idx += 1
        
    # Certifications Section
    cert_start_row = row_idx + 2
    ws5.cell(row=cert_start_row, column=1, value="Certifications").font = section_title_font
    
    headers_cert = ["Certification Name", "Issuing Organization", "Date"]
    for i, h in enumerate(headers_cert, 1):
        ws5.cell(row=cert_start_row + 1, column=i, value=h)
    style_header(ws5, 1, len(headers_cert), row=cert_start_row + 1)
    
    row_idx = cert_start_row + 2
    for cert in resume.certifications:
        c1 = ws5.cell(row=row_idx, column=1, value=cert.name)
        c2 = ws5.cell(row=row_idx, column=2, value=cert.issuer)
        c3 = ws5.cell(row=row_idx, column=3, value=cert.date)
        
        for cell in [c1, c2, c3]:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_left
            
        c1.font = bold_font
        ws5.row_dimensions[row_idx].height = 24
        row_idx += 1
        
    auto_fit_columns(ws5, min_width=18, max_width=60)

    # ================= Sheet 6: Languages =================
    ws6 = wb.create_sheet(title="Languages")
    ws6.views.sheetView[0].showGridLines = True
    
    headers = ["Language", "Proficiency Level"]
    for i, h in enumerate(headers, 1):
        ws6.cell(row=1, column=i, value=h)
    style_header(ws6, 1, len(headers))
    
    row_idx = 2
    for lang in resume.languages:
        c1 = ws6.cell(row=row_idx, column=1, value=lang.name)
        c2 = ws6.cell(row=row_idx, column=2, value=lang.proficiency)
        
        c1.font = bold_font
        c1.border = thin_border
        c1.alignment = align_left
        
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = align_left
        
        ws6.row_dimensions[row_idx].height = 24
        row_idx += 1
        
    auto_fit_columns(ws6, min_width=20, max_width=30)
    
    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


# ================= HTML Resume templates (resume.io style) =================

def render_html_resume(resume: ResumeData, template_name: str = "Dublin") -> str:
    """
    Renders the ResumeData into a beautiful HTML/CSS resume.
    Supported templates: Dublin, Toronto, Stockholm
    """
    # Build sections
    
    # Work Experience HTML
    experience_html = ""
    for exp in resume.experience:
        bullets = "".join([f"<li>{item}</li>" for item in exp.description])
        experience_html += f"""
        <div class="resume-item">
            <div class="item-header">
                <div>
                    <h3 class="item-title">{exp.position}</h3>
                    <div class="item-subtitle">{exp.company} | {exp.location}</div>
                </div>
                <div class="item-date">{exp.start_date} - {exp.end_date}</div>
            </div>
            <ul class="item-details">
                {bullets}
            </ul>
        </div>
        """
        
    # Education HTML
    education_html = ""
    for edu in resume.education:
        details = f"<p class='edu-details'>{edu.description}</p>" if edu.description else ""
        education_html += f"""
        <div class="resume-item">
            <div class="item-header">
                <div>
                    <h3 class="item-title">{edu.degree}</h3>
                    <div class="item-subtitle">{edu.institution} | {edu.location}</div>
                </div>
                <div class="item-date">{edu.start_date} - {edu.end_date}</div>
            </div>
            {details}
        </div>
        """
        
    # Skills HTML
    skills_html = ""
    for group in resume.skills:
        skills_pills = "".join([f"<span class='skill-pill'>{s}</span>" for s in group.skills])
        skills_html += f"""
        <div class="skill-category-group">
            <h4 class="skill-group-title">{group.category}</h4>
            <div class="skills-container">{skills_pills}</div>
        </div>
        """
        
    # Projects HTML
    projects_html = ""
    for p in resume.projects:
        link_html = f" &bull; <a href='{p.link}' target='_blank' class='item-link'>Project Link</a>" if p.link else ""
        role_html = f" ({p.role})" if p.role else ""
        projects_html += f"""
        <div class="resume-item">
            <div class="item-header">
                <h3 class="item-title">{p.name}{role_html}{link_html}</h3>
            </div>
            <p class="item-text">{p.description}</p>
        </div>
        """
        
    # Certifications HTML
    certs_html = ""
    for c in resume.certifications:
        certs_html += f"""
        <div class="cert-item">
            <strong>{c.name}</strong> &ndash; {c.issuer} ({c.date})
        </div>
        """
        
    # Languages HTML
    lang_html = ""
    for l in resume.languages:
        lang_html += f"""
        <div class="lang-item">
            <span class="lang-name">{l.name}</span>: <span class="lang-level">{l.proficiency}</span>
        </div>
        """

    # Template Styles
    if template_name.lower() == "dublin": # Two column layout with right sidebar
        css_styles = """
        :root {
            --primary: #1e293b;
            --accent: #00b289;
            --text-dark: #1e293b;
            --text-light: #475569;
            --text-muted: #94a3b8;
            --border-color: #ebebeb;
        }
        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            color: var(--text-dark);
            line-height: 1.5;
            margin: 0;
            padding: 0;
            background-color: #fff;
        }
        .resume-container {
            width: 210mm;
            min-height: 297mm;
            margin: 0 auto;
            padding: 50px 45px;
            box-sizing: border-box;
            background-color: #fff;
            border: 1px solid var(--border-color);
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
        }
        .header {
            display: flex;
            align-items: center;
            gap: 20px;
            padding-bottom: 25px;
            border-bottom: 1.5px solid var(--border-color);
            margin-bottom: 25px;
        }
        .profile-photo {
            width: 80px;
            height: 80px;
            border-radius: 8px;
            object-fit: cover;
            border: 1px solid var(--border-color);
        }
        .header-text {
            display: flex;
            flex-direction: column;
        }
        .header-text h1 {
            font-size: 28px;
            font-weight: 700;
            color: var(--primary);
            margin: 0 0 4px 0;
            letter-spacing: -0.5px;
        }
        .header-text h2 {
            font-size: 13.5px;
            font-weight: 500;
            color: var(--text-light);
            margin: 0;
            text-transform: none;
            letter-spacing: 0.5px;
        }
        .resume-body {
            display: flex;
            gap: 45px;
        }
        .main-content {
            width: 65%;
            display: flex;
            flex-direction: column;
            gap: 22px;
        }
        .sidebar {
            width: 35%;
            display: flex;
            flex-direction: column;
            gap: 22px;
        }
        .section {
            margin-bottom: 5px;
        }
        .section-title {
            display: flex;
            align-items: center;
            font-size: 15px;
            font-weight: 700;
            color: var(--primary);
            margin-top: 0;
            margin-bottom: 12px;
            border-bottom: none;
            padding-bottom: 0;
        }
        .section-title svg {
            width: 16px;
            height: 16px;
            fill: var(--accent);
            margin-right: 8px;
            vertical-align: middle;
        }
        .sidebar-section {
            margin-bottom: 5px;
        }
        .sidebar-section-title {
            font-size: 13.5px;
            font-weight: 700;
            color: var(--primary);
            margin-top: 0;
            margin-bottom: 12px;
            border-bottom: none;
            padding-bottom: 0;
        }
        .contact-info {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        .contact-item {
            display: flex;
            flex-direction: column;
            font-size: 11.5px;
        }
        .contact-item .label {
            font-size: 9px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 2px;
        }
        .contact-item .value {
            color: var(--text-dark);
            font-size: 11.5px;
            line-height: 1.4;
            white-space: pre-line;
        }
        .contact-item a {
            color: var(--accent);
            text-decoration: none;
            font-weight: 500;
        }
        .contact-item a:hover {
            text-decoration: underline;
        }
        .summary-text {
            font-size: 12px;
            color: var(--text-light);
            margin: 0;
            line-height: 1.5;
            text-align: justify;
        }
        .resume-item {
            margin-bottom: 14px;
        }
        .resume-item:last-child {
            margin-bottom: 0;
        }
        .item-header {
            margin-bottom: 3px;
        }
        .item-title {
            font-size: 12.5px;
            font-weight: 700;
            color: var(--primary);
            margin: 0;
            line-height: 1.4;
        }
        .item-subtitle {
            font-size: 11.5px;
            color: var(--text-light);
            font-weight: 500;
        }
        .item-date {
            font-size: 10.5px;
            color: var(--text-muted);
            margin-top: 1px;
            margin-bottom: 3px;
        }
        .item-details {
            margin: 0;
            padding-left: 15px;
            font-size: 11.5px;
            color: var(--text-light);
            line-height: 1.45;
        }
        .item-details li {
            margin-bottom: 3px;
        }
        .item-text {
            font-size: 11.5px;
            color: var(--text-light);
            margin: 3px 0 0 0;
            line-height: 1.4;
        }
        .edu-details {
            font-size: 11.5px;
            color: var(--text-light);
            margin: 3px 0 0 0;
            line-height: 1.4;
        }
        .skill-category-group {
            margin-bottom: 12px;
        }
        .skill-item {
            margin-bottom: 8px;
        }
        .skill-name {
            font-size: 11.5px;
            font-weight: 500;
            color: var(--text-dark);
        }
        .skill-bar {
            width: 100%;
            height: 4px;
            background-color: var(--border-color);
            border-radius: 2px;
            margin-top: 3px;
        }
        .skill-bar-fill {
            height: 100%;
            background-color: var(--accent);
            border-radius: 2px;
        }
        .lang-item {
            font-size: 11.5px;
            color: var(--text-light);
            margin-bottom: 4px;
            line-height: 1.4;
        }
        .lang-name {
            font-weight: 600;
            color: var(--text-dark);
        }
        .lang-level {
            color: var(--text-light);
        }
        .cert-item {
            font-size: 11.5px;
            color: var(--text-light);
            margin-bottom: 6px;
            line-height: 1.4;
        }
        .item-link {
            color: var(--accent);
            text-decoration: none;
            font-size: 10.5px;
            font-weight: 500;
        }
        .item-link:hover {
            text-decoration: underline;
        }
        @media print {
            body {
                background-color: #fff;
            }
            .resume-container {
                box-shadow: none;
                border: none;
                margin: 0;
                padding: 0;
                width: 100%;
                height: 100%;
            }
        }
        """

        # Section SVG Icons
        profile_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M12 12c2.21 0 4-1.79 4-4s-1.79-4-4-4-4 1.79-4 4 1.79 4 4 4zm0 2c-2.67 0-8 1.34-8 4v2h16v-2c0-2.66-5.33-4-8-4z"/></svg>"""
        work_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M20 6h-4V4c0-1.11-.89-2-2-2h-4c-1.11 0-2 .89-2 2v2H4c-1.11 0-1.99.89-1.99 2L2 19c0 1.11.89 2 2 2h16c1.11 0 2-.89 2-2V8c0-1.11-.89-2-2-2zm-6 0h-4V4h4v2z"/></svg>"""
        edu_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M5 13.18v4L12 21l7-3.82v-4L12 17l-7-3.82zM12 3L1 9l11 6 9-4.91v3.51h2V9L12 3z"/></svg>"""
        project_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M12 2s3 5.5 3 8c0 1.66-1.34 3-3 3s-3-1.34-3-3c0-2.5 3-8 3-8zm-3 12c0 2.21 1.79 4 4 4s4-1.79 4-4H9zm4 6h-2v2h2v-2z"/></svg>"""
        cert_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M12 22c5.523 0 10-4.477 10-10S17.523 2 12 2 2 6.477 2 12s4.477 10 10 10zm1-15h-2v6h2V7zm0 8h-2v2h2v-2z"/></svg>"""
        accomplishment_icon = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><path d="M18 2H6c-1.1 0-2 .9-2 2v3c0 2.24 1.5 4.15 3.58 4.75C8.26 13.06 10 14.33 11 15.15V19H9c-.55 0-1 .45-1 1s.45 1 1 1h6c.55 0 1-.45 1-1s-.45-1-1-1h-2v-3.85c1-.82 2.74-2.09 3.42-3.4C18.5 11.15 20 9.24 20 7V4c0-1.1-.9-2-2-2zm-6 5V4h2v3c0 .85-.19 1.63-.5 2.28C6.67 8.65 6 7.9 6 7zm12 0c0 .9-.67 1.65-1.5 2.28-.31-.65-.5-1.43-.5-2.28V4h2v3z"/></svg>"""

        # Setup links for Dublin
        links_list = []
        if resume.personal_info.linkedin:
            links_list.append(f'<div class="contact-item"><a href="{resume.personal_info.linkedin}" target="_blank">LinkedIn</a></div>')
        if resume.personal_info.github:
            links_list.append(f'<div class="contact-item"><a href="{resume.personal_info.github}" target="_blank">GitHub</a></div>')
        if resume.personal_info.website:
            links_list.append(f'<div class="contact-item"><a href="{resume.personal_info.website}" target="_blank">Portfolio</a></div>')
        links_html_dublin = "".join(links_list)

        # Override skills_html for Dublin to use progress bars
        skills_html_dublin = ""
        for idx_cat, group in enumerate(resume.skills):
            skills_html_dublin += f'<div class="skill-category-group"><h4 style="font-size:11px; color:var(--text-dark); margin: 0 0 6px 0;">{group.category}</h4>'
            for idx_s, s in enumerate(group.skills):
                # Stagger progress bars for Dublin layout
                level = max(60, 95 - (idx_s * 8))
                skills_html_dublin += f"""
                <div class="skill-item">
                    <div class="skill-name">{s}</div>
                    <div class="skill-bar"><div class="skill-bar-fill" style="width: {level}%;"></div></div>
                </div>
                """
            skills_html_dublin += '</div>'

        # Override lang_html for Dublin
        lang_html_dublin = ""
        for l in resume.languages:
            lang_html_dublin += f"""
            <div class="lang-item">
                <span class="lang-name">{l.name}</span> <span class="lang-level">({l.proficiency})</span>
            </div>
            """

        # Accomplishments HTML
        accomplishments_html = ""
        if hasattr(resume, 'accomplishments') and resume.accomplishments:
            acc_bullets = "".join([f"<li>{item}</li>" for item in resume.accomplishments])
            accomplishments_html = f"""
            <div class="section">
                <h3 class="section-title">{accomplishment_icon}Accomplishments</h3>
                <ul class="item-details" style="margin-top:5px;">
                    {acc_bullets}
                </ul>
            </div>
            """

        html_output = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{resume.personal_info.full_name} - Resume</title>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
            <style>
                {css_styles}
            </style>
        </head>
        <body>
            <div class="resume-container">
                <div class="header">
                    {f'<img class="profile-photo" src="{resume.personal_info.photo_base64}" alt="Profile Photo">' if resume.personal_info.photo_base64 else ''}
                    <div class="header-text">
                        <h1>{resume.personal_info.full_name}</h1>
                        <h2>{resume.personal_info.professional_title}</h2>
                        {f'<div class="power-statement" style="font-style: italic; font-size: 11px; color: var(--text-light); margin-top: 5px; max-width: 500px; font-weight: 500;">"{resume.personal_info.power_statement}"</div>' if hasattr(resume.personal_info, "power_statement") and resume.personal_info.power_statement else ''}
                    </div>
                </div>
                
                <div class="resume-body">
                    <div class="main-content">
                        {f'<div class="section"><h3 class="section-title">{profile_icon}Profile</h3><p class="summary-text">{resume.summary}</p></div>' if resume.summary else ''}
                        
                        {f'<div class="section"><h3 class="section-title">{work_icon}Employment History</h3>{experience_html}</div>' if resume.experience else ''}
                        
                        {accomplishments_html}
                        
                        {f'<div class="section"><h3 class="section-title">{edu_icon}Education</h3>{education_html}</div>' if resume.education else ''}
                        
                        {f'<div class="section"><h3 class="section-title">{project_icon}Key Projects</h3>{projects_html}</div>' if resume.projects else ''}
                        
                        {f'<div class="section"><h3 class="section-title">{cert_icon}Certifications</h3>{certs_html}</div>' if resume.certifications else ''}
                    </div>
                    
                    <div class="sidebar">
                        <div class="sidebar-section">
                            <h4 class="sidebar-section-title">Details</h4>
                            <div class="contact-info">
                                <div class="contact-item">
                                    <span class="label">Location</span>
                                    <span class="value">{resume.personal_info.location}</span>
                                </div>
                                <div class="contact-item">
                                    <span class="label">Phone</span>
                                    <span class="value">{resume.personal_info.phone}</span>
                                </div>
                                <div class="contact-item">
                                    <span class="label">Email</span>
                                    <span class="value"><a href="mailto:{resume.personal_info.email}">{resume.personal_info.email}</a></span>
                                </div>
                                {f'<div class="contact-item"><span class="label">Date / Place of Birth</span><span class="value">{resume.personal_info.birth_info}</span></div>' if resume.personal_info.birth_info else ''}
                                {f'<div class="contact-item"><span class="label">Nationality</span><span class="value">{resume.personal_info.nationality}</span></div>' if resume.personal_info.nationality else ''}
                            </div>
                        </div>
                        
                        {f'<div class="sidebar-section"><h4 class="sidebar-section-title">Links</h4><div class="contact-info">{links_html_dublin}</div></div>' if links_html_dublin else ''}
                        
                        {f'<div class="sidebar-section"><h4 class="sidebar-section-title">Skills</h4>{skills_html_dublin}</div>' if resume.skills else ''}
                        
                        {f'<div class="sidebar-section"><h4 class="sidebar-section-title">Languages</h4>{lang_html_dublin}</div>' if resume.languages else ''}
                        
                        {f'<div class="sidebar-section"><h4 class="sidebar-section-title">Hobbies</h4><p style="font-size: 11.5px; color: var(--text-light); margin: 0; line-height: 1.4;">{resume.personal_info.hobbies}</p></div>' if resume.personal_info.hobbies else ''}
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        return html_output

    elif template_name.lower() == "toronto": # Modern layout with colored accent bar
        css_styles = """
        :root {
            --accent: #3182CE;
            --accent-dark: #2B6CB0;
            --text-dark: #1A202C;
            --text-light: #4A5568;
            --border-color: #E2E8F0;
        }
        body {
            font-family: 'Outfit', sans-serif;
            color: var(--text-dark);
            line-height: 1.6;
            margin: 0;
            padding: 0;
            background-color: #fff;
        }
        .resume-container {
            width: 210mm;
            min-height: 297mm;
            margin: 0 auto;
            padding: 40px;
            border: 1px solid var(--border-color);
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
            box-sizing: border-box;
            position: relative;
        }
        /* Top Accent Bar */
        .resume-container::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 8px;
            background-color: var(--accent);
        }
        .header {
            display: flex;
            justify-content: space-between;
            align-items: flex-end;
            border-bottom: 2px solid var(--border-color);
            padding-bottom: 20px;
            margin-bottom: 25px;
        }
        .header-left h1 {
            font-size: 32px;
            color: var(--text-dark);
            margin: 0 0 5px 0;
            font-weight: 800;
            letter-spacing: -0.5px;
        }
        .header-left h2 {
            font-size: 16px;
            color: var(--accent);
            margin: 0;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 1px;
        }
        .header-right {
            text-align: right;
            font-size: 11px;
            color: var(--text-light);
            display: flex;
            flex-direction: column;
            gap: 2px;
        }
        .header-right a {
            color: var(--accent-dark);
            text-decoration: none;
        }
        .section {
            margin-bottom: 25px;
        }
        .section-title {
            font-size: 16px;
            color: var(--text-dark);
            text-transform: uppercase;
            letter-spacing: 1px;
            border-bottom: 1px solid var(--accent);
            padding-bottom: 4px;
            margin-top: 0;
            margin-bottom: 15px;
            font-weight: 700;
        }
        .summary-text {
            font-size: 13px;
            color: var(--text-light);
            margin: 0;
        }
        .resume-item {
            margin-bottom: 15px;
        }
        .item-header {
            display: flex;
            justify-content: space-between;
            align-items: baseline;
            margin-bottom: 4px;
        }
        .item-title {
            font-size: 14px;
            font-weight: 700;
            color: var(--text-dark);
            margin: 0;
        }
        .item-subtitle {
            font-size: 13px;
            color: var(--accent-dark);
            font-weight: 600;
        }
        .item-date {
            font-size: 12px;
            color: var(--text-light);
            font-weight: 500;
        }
        .item-details {
            margin: 4px 0 0 0;
            padding-left: 20px;
            font-size: 12.5px;
            color: var(--text-light);
        }
        .item-details li {
            margin-bottom: 3px;
        }
        .item-text {
            font-size: 12.5px;
            color: var(--text-light);
            margin: 4px 0 0 0;
        }
        .skills-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 15px;
        }
        .skill-category-group {
            margin-bottom: 0;
        }
        .skill-group-title {
            font-size: 12px;
            font-weight: 700;
            margin: 0 0 5px 0;
            color: var(--text-dark);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .skills-container {
            display: flex;
            flex-wrap: wrap;
            gap: 5px;
        }
        .skill-pill {
            background-color: #EDF2F7;
            color: var(--text-light);
            font-size: 11px;
            padding: 2px 7px;
            border-radius: 3px;
        }
        .grid-2col {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
        }
        .cert-item {
            font-size: 12.5px;
            color: var(--text-light);
            margin-bottom: 6px;
        }
        .lang-item {
            font-size: 12.5px;
            color: var(--text-light);
            margin-bottom: 6px;
        }
        .lang-name {
            font-weight: 600;
        }
        .item-link {
            color: var(--accent);
            text-decoration: none;
        }
        @media print {
            .resume-container {
                box-shadow: none;
                border: none;
                margin: 0;
                padding: 20px;
                width: 100%;
                height: 100%;
            }
        }
        """
        
        # Format contact line info
        contact_line = f"<span>{resume.personal_info.phone}</span> &bull; <span>{resume.personal_info.email}</span>"
        if resume.personal_info.location:
            contact_line += f" &bull; <span>{resume.personal_info.location}</span>"
            
        extra_info_list = []
        if resume.personal_info.birth_info:
            extra_info_list.append(f"Born: {resume.personal_info.birth_info}")
        if resume.personal_info.nationality:
            extra_info_list.append(f"Nationality: {resume.personal_info.nationality}")
        extra_info_line = " &bull; ".join(extra_info_list) if extra_info_list else ""
            
        links_list = []
        if resume.personal_info.linkedin:
            links_list.append(f'<a href="{resume.personal_info.linkedin}" target="_blank">LinkedIn</a>')
        if resume.personal_info.github:
            links_list.append(f'<a href="{resume.personal_info.github}" target="_blank">GitHub</a>')
        if resume.personal_info.website:
            links_list.append(f'<a href="{resume.personal_info.website}" target="_blank">Website</a>')
        links_line = " &bull; ".join(links_list)

        html_output = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{resume.personal_info.full_name} - Resume</title>
            <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800&display=swap" rel="stylesheet">
            <style>
                {css_styles}
            </style>
        </head>
        <body>
            <div class="resume-container">
                <div class="header">
                    <div style="display: flex; align-items: center; gap: 15px;">
                        {f'<img style="width: 75px; height: 75px; border-radius: 50%; object-fit: cover; border: 2px solid var(--accent); display: block;" src="{resume.personal_info.photo_base64}" alt="Profile Photo">' if resume.personal_info.photo_base64 else ''}
                        <div class="header-left">
                            <h1>{resume.personal_info.full_name}</h1>
                            <h2>{resume.personal_info.professional_title}</h2>
                            {f'<div style="font-style: italic; font-size: 11px; color: var(--text-light); margin-top: 4px;">"{resume.personal_info.power_statement}"</div>' if getattr(resume.personal_info, "power_statement", None) else ''}
                        </div>
                    </div>
                    <div class="header-right">
                        <div>{contact_line}</div>
                        {f'<div>{extra_info_line}</div>' if extra_info_line else ''}
                        <div>{links_line}</div>
                    </div>
                </div>
                
                <div class="section">
                    <h3 class="section-title">Professional Summary</h3>
                    <p class="summary-text">{resume.summary}</p>
                </div>
                
                {f'<div class="section"><h3 class="section-title">Professional Experience</h3>{experience_html}</div>' if resume.experience else ''}
                
                {f'<div class="section"><h3 class="section-title">Accomplishments</h3><ul class="item-details" style="margin-top: 5px; padding-left: 20px; font-size: 12.5px; color: var(--text-light);">{"".join([f"<li style=\'margin-bottom:3px;\'>{item}</li>" for item in resume.accomplishments])}</ul></div>' if getattr(resume, "accomplishments", None) else ''}
                
                {f'<div class="section"><h3 class="section-title">Education</h3>{education_html}</div>' if resume.education else ''}
                
                {f'<div class="section"><h3 class="section-title">Key Projects</h3>{projects_html}</div>' if resume.projects else ''}
                
                <div class="grid-2col">
                    {f'<div class="section"><h3 class="section-title">Skills</h3><div class="skills-grid">{skills_html}</div></div>' if resume.skills else ''}
                    
                    <div class="section">
                        {f'<h3 class="section-title">Certifications</h3>{certs_html}' if resume.certifications else ''}
                        {f'<h3 class="section-title" style="margin-top:15px;">Languages</h3>{lang_html}' if resume.languages else ''}
                        {f'<h3 class="section-title" style="margin-top:15px;">Hobbies</h3><p style="font-size: 12.5px; color: var(--text-light); margin: 0; line-height: 1.5;">{resume.personal_info.hobbies}</p>' if resume.personal_info.hobbies else ''}
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        return html_output

    else: # Stockholm Template (Minimalist Elegant)
        css_styles = """
        :root {
            --text-dark: #000000;
            --text-muted: #555555;
            --border-color: #000000;
        }
        body {
            font-family: 'Playfair Display', Georgia, serif;
            color: var(--text-dark);
            line-height: 1.5;
            margin: 0;
            padding: 0;
            background-color: #fff;
        }
        .resume-container {
            width: 210mm;
            min-height: 297mm;
            margin: 0 auto;
            padding: 50px;
            border: 1px solid #ddd;
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);
            box-sizing: border-box;
        }
        .header {
            text-align: center;
            margin-bottom: 25px;
        }
        .header h1 {
            font-size: 36px;
            color: var(--text-dark);
            margin: 0 0 5px 0;
            font-weight: 400;
            text-transform: uppercase;
            letter-spacing: 2px;
        }
        .header h2 {
            font-size: 14px;
            color: var(--text-muted);
            margin: 0 0 15px 0;
            font-weight: 400;
            text-transform: uppercase;
            letter-spacing: 3px;
        }
        .contact-bar {
            display: flex;
            justify-content: center;
            flex-wrap: wrap;
            gap: 12px;
            font-size: 11px;
            color: var(--text-muted);
            border-top: 1px solid var(--border-color);
            border-bottom: 1px solid var(--border-color);
            padding: 8px 0;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-family: 'Inter', sans-serif;
        }
        .contact-bar a {
            color: var(--text-dark);
            text-decoration: none;
            font-weight: 500;
        }
        .section {
            margin-top: 25px;
        }
        .section-title {
            font-size: 13px;
            color: var(--text-dark);
            text-transform: uppercase;
            letter-spacing: 2px;
            border-bottom: 1px double var(--border-color);
            padding-bottom: 3px;
            margin-top: 0;
            margin-bottom: 15px;
            font-weight: 600;
            font-family: 'Inter', sans-serif;
        }
        .summary-text {
            font-size: 13.5px;
            color: #333;
            margin: 0;
            text-align: justify;
        }
        .resume-item {
            margin-bottom: 18px;
        }
        .item-header {
            display: flex;
            justify-content: space-between;
            align-items: baseline;
            margin-bottom: 4px;
        }
        .item-title {
            font-size: 14px;
            font-weight: 600;
            color: var(--text-dark);
            margin: 0;
            font-family: 'Inter', sans-serif;
        }
        .item-subtitle {
            font-size: 13px;
            color: var(--text-muted);
            font-style: italic;
        }
        .item-date {
            font-size: 12px;
            color: var(--text-muted);
            font-family: 'Inter', sans-serif;
        }
        .item-details {
            margin: 4px 0 0 0;
            padding-left: 20px;
            font-size: 13px;
            color: #333;
        }
        .item-details li {
            margin-bottom: 3px;
        }
        .item-text {
            font-size: 13px;
            color: #333;
            margin: 4px 0 0 0;
        }
        .skills-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 15px;
        }
        .skill-category-group {
            margin-bottom: 0;
        }
        .skill-group-title {
            font-size: 11px;
            font-weight: 700;
            margin: 0 0 4px 0;
            color: var(--text-dark);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-family: 'Inter', sans-serif;
        }
        .skills-container {
            display: flex;
            flex-wrap: wrap;
            gap: 4px;
        }
        .skill-pill {
            font-size: 11.5px;
            color: #333;
        }
        .skill-pill::after {
            content: ',';
        }
        .skill-pill:last-child::after {
            content: '';
        }
        .cert-item {
            font-size: 13px;
            color: #333;
            margin-bottom: 6px;
        }
        .lang-item {
            font-size: 13px;
            color: #333;
            margin-bottom: 6px;
        }
        .lang-name {
            font-weight: 600;
        }
        .item-link {
            color: var(--text-dark);
            text-decoration: underline;
        }
        @media print {
            .resume-container {
                box-shadow: none;
                border: none;
                margin: 0;
                padding: 10px;
                width: 100%;
                height: 100%;
            }
        }
        """
        
        # Setup links
        contact_links = []
        contact_links.append(resume.personal_info.phone)
        contact_links.append(f'<a href="mailto:{resume.personal_info.email}">{resume.personal_info.email}</a>')
        contact_links.append(resume.personal_info.location)
        if resume.personal_info.birth_info:
            contact_links.append(f"Born: {resume.personal_info.birth_info}")
        if resume.personal_info.nationality:
            contact_links.append(f"Nationality: {resume.personal_info.nationality}")
        if resume.personal_info.linkedin:
            contact_links.append(f'<a href="{resume.personal_info.linkedin}" target="_blank">LinkedIn</a>')
        if resume.personal_info.github:
            contact_links.append(f'<a href="{resume.personal_info.github}" target="_blank">GitHub</a>')
        if resume.personal_info.website:
            contact_links.append(f'<a href="{resume.personal_info.website}" target="_blank">Website</a>')
        
        contact_bar_content = " &bull; ".join(contact_links)

        # Add profile photo styles to CSS if photo exists
        photo_css = """
        .profile-photo-container {
            text-align: center;
            margin-bottom: 20px;
        }
        .profile-photo {
            width: 100px;
            height: 100px;
            border-radius: 50%;
            object-fit: cover;
            border: 1px solid var(--border-color);
            display: inline-block;
        }
        """
        css_styles_with_photo = css_styles + photo_css

        html_output = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{resume.personal_info.full_name} - Resume</title>
            <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400;0,600;1,400&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
            <style>
                {css_styles_with_photo}
            </style>
        </head>
        <body>
            <div class="resume-container">
                <div class="header">
                    {f'<div class="profile-photo-container"><img class="profile-photo" src="{resume.personal_info.photo_base64}" alt="Profile Photo"></div>' if resume.personal_info.photo_base64 else ''}
                    <h1>{resume.personal_info.full_name}</h1>
                    <h2>{resume.personal_info.professional_title}</h2>
                    {f'<div style="font-style: italic; font-size: 11.5px; color: var(--text-muted); margin-bottom: 10px;">"{resume.personal_info.power_statement}"</div>' if getattr(resume.personal_info, "power_statement", None) else ''}
                    <div class="contact-bar">
                        {contact_bar_content}
                    </div>
                </div>
                
                <div class="section">
                    <h3 class="section-title">Summary</h3>
                    <p class="summary-text">{resume.summary}</p>
                </div>
                
                {f'<div class="section"><h3 class="section-title">Experience</h3>{experience_html}</div>' if resume.experience else ''}
                
                {f'<div class="section"><h3 class="section-title">Accomplishments</h3><ul class="item-details" style="margin-top: 5px; padding-left: 20px; font-size: 13px; color: #333;">{"".join([f"<li style=\'margin-bottom:3px;\'>{item}</li>" for item in resume.accomplishments])}</ul></div>' if getattr(resume, "accomplishments", None) else ''}
                
                {f'<div class="section"><h3 class="section-title">Education</h3>{education_html}</div>' if resume.education else ''}
                
                {f'<div class="section"><h3 class="section-title">Projects</h3>{projects_html}</div>' if resume.projects else ''}
                
                {f'<div class="section"><h3 class="section-title">Skills</h3><div class="skills-grid">{skills_html}</div></div>' if resume.skills else ''}
                
                {f'<div class="section"><h3 class="section-title">Certifications</h3>{certs_html}</div>' if resume.certifications else ''}
                
                {f'<div class="section"><h3 class="section-title">Languages</h3>{lang_html}</div>' if resume.languages else ''}
                
                {f'<div class="section"><h3 class="section-title">Hobbies</h3><p style="font-size: 13.5px; color: #333; margin: 0; line-height: 1.5;">{resume.personal_info.hobbies}</p></div>' if resume.personal_info.hobbies else ''}
            </div>
        </body>
        </html>
        """
        return html_output
